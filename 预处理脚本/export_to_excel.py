import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def load_comparison_results(file_path):
    """加载对比结果"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def safe_join_list(data_list):
    """安全地连接列表，处理可能的非字符串元素"""
    if not data_list:
        return ""
    
    str_list = []
    for item in data_list:
        if isinstance(item, (dict, list)):
            # 如果是字典或列表，保留JSON格式
            str_list.append(json.dumps(item, ensure_ascii=False))
        else:
            str_list.append(str(item))
    
    return " | ".join(str_list)

def prepare_excel_data(comparison_results):
    """准备Excel数据"""
    excel_data = []
    
    for result in comparison_results:
        gso_chapter = result["gso_chapter"]
        ece_candidates = result["ece_candidates"]
        match_info = result["match_info"]
        
        # GSO章节信息
        gso_info = {
            "GSO_Chapter_ID": match_info["chapter_id"],
            "GSO_Topic_Keywords": safe_join_list(gso_chapter.get("topic_keywords", [])),
            "GSO_Context_Keywords": safe_join_list(gso_chapter.get("context_keywords", [])),
            "GSO_Parameters": json.dumps(gso_chapter.get("parameters", []), ensure_ascii=False, indent=2) if gso_chapter.get("parameters") else "",
            "GSO_Table_Headers": safe_join_list(gso_chapter.get("table_headers", [])),
            "Ground_Truth": ""  # 新增列，用于填写正确答案
        }
        
        # 如果没有候选项，添加一行空数据
        if not ece_candidates:
            row = gso_info.copy()
            row.update({
                "Candidate_Rank": "",
                "Match_Score": "",
                "ECE_Chapter_Path": "",
                "ECE_Topic_Keywords": "",
                "ECE_Context_Keywords": "",
                "ECE_Parameters": "",
                "ECE_Table_Headers": ""
            })
            excel_data.append(row)
        else:
            # 为每个候选项添加一行
            for i, candidate in enumerate(ece_candidates):
                row = gso_info.copy() if i == 0 else {key: "" for key in gso_info.keys()}
                
                comparison_data = candidate.get("comparison_data", {})
                if comparison_data is None:
                    comparison_data = {}
                
                # 处理ECE章节路径显示
                chapter_path = candidate.get("chapter_path", [])
                if len(chapter_path) >= 3:
                    file_name, section_name, chapter_id = chapter_path[0], chapter_path[1], chapter_path[2]
                    if section_name == "MAIN":
                        # 如果是MAIN部分，只显示章节号
                        ece_chapter_display = chapter_id
                    else:
                        # 如果是Appendix等其他部分，显示section和章节号
                        ece_chapter_display = f"{section_name} > {chapter_id}"
                else:
                    ece_chapter_display = " > ".join(chapter_path)
                
                row.update({
                    "Candidate_Rank": i + 1,
                    "Match_Score": f"{candidate.get('score', 0):.4f}",
                    "ECE_Chapter_Path": ece_chapter_display,
                    "ECE_Topic_Keywords": safe_join_list(comparison_data.get("topic_keywords", [])),
                    "ECE_Context_Keywords": safe_join_list(comparison_data.get("context_keywords", [])),
                    "ECE_Parameters": json.dumps(comparison_data.get("parameters", []), ensure_ascii=False, indent=2) if comparison_data.get("parameters") else "",
                    "ECE_Table_Headers": safe_join_list(comparison_data.get("table_headers", [])),
                    "Ground_Truth": ""  # 空列供后续填写
                })
                
                excel_data.append(row)
    
    return excel_data

def style_excel_workbook(workbook, worksheet, total_rows):
    """设置Excel样式"""
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    gso_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    ece_fill = PatternFill(start_color="F0F8E7", end_color="F0F8E7", fill_type="solid")
    ground_truth_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # 浅橙色
    
    high_score_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 浅绿色
    medium_score_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # 浅黄色
    low_score_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")  # 浅红色
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 设置列宽
    column_widths = {
        'A': 12,  # GSO_Chapter_ID
        'B': 20,  # GSO_Topic_Keywords
        'C': 20,  # GSO_Context_Keywords
        'D': 20,  # GSO_Parameters (加宽以容纳JSON)
        'E': 10,  # GSO_Table_Headers
        'F': 12,  # Ground_Truth
        'G': 8,   # Candidate_Rank
        'H': 12,  # Match_Score
        'I': 15,  # ECE_Chapter_Path
        'J': 30,  # ECE_Topic_Keywords
        'K': 25,  # ECE_Context_Keywords
        'L': 50,  # ECE_Parameters (加宽以容纳JSON)
        'M': 50   # ECE_Table_Headers
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # 设置表头样式
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # 设置数据行样式
    for row_num in range(2, total_rows + 2):
        for col_num in range(1, 14):  # A到M列
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Ground Truth列（F列）使用橙色背景
            if col_num == 6:
                cell.fill = ground_truth_fill
            # GSO列（A-E）使用蓝色背景
            elif col_num <= 5:
                cell.fill = gso_fill
            # ECE列使用绿色背景
            else:
                cell.fill = ece_fill

def merge_gso_cells(worksheet, excel_data):
    """合并GSO相关的单元格"""
    current_gso_chapter = None
    start_row = 2
    
    for i, row_data in enumerate(excel_data, start=2):
        gso_chapter_id = row_data["GSO_Chapter_ID"]
        
        if gso_chapter_id and gso_chapter_id != current_gso_chapter:
            # 如果不是第一个章节，合并前一个章节的单元格
            if current_gso_chapter is not None and i > start_row:
                for col in range(1, 7):  # A到F列（包括Ground_Truth）
                    if i - start_row > 1:  # 只有当有多行时才合并
                        worksheet.merge_cells(
                            start_row=start_row, start_column=col,
                            end_row=i-1, end_column=col
                        )
                        # 设置合并单元格的对齐方式
                        cell = worksheet.cell(row=start_row, column=col)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            current_gso_chapter = gso_chapter_id
            start_row = i
    
    # 处理最后一个章节
    if current_gso_chapter is not None:
        end_row = len(excel_data) + 1
        if end_row > start_row:
            for col in range(1, 7):  # A到F列（包括Ground_Truth）
                if end_row - start_row > 1:
                    worksheet.merge_cells(
                        start_row=start_row, start_column=col,
                        end_row=end_row, end_column=col
                    )
                    cell = worksheet.cell(row=start_row, column=col)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def create_excel_report(comparison_results, output_file):
    """创建Excel报告"""
    
    # 准备数据
    excel_data = prepare_excel_data(comparison_results)
    
    # 创建DataFrame
    df = pd.DataFrame(excel_data)
    
    # 创建工作簿
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "GSO-ECE Mapping Comparison"
    
    # 添加数据到工作表
    for r in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(r)
    
    # 设置样式
    style_excel_workbook(workbook, worksheet, len(excel_data))
    
    # 合并GSO单元格
    merge_gso_cells(worksheet, excel_data)
    
    # 冻结首行
    worksheet.freeze_panes = "A2"
    
    # 保存文件
    workbook.save(output_file)
    print(f"Excel报告已保存到: {output_file}")

def create_summary_sheet(workbook, comparison_results):
    """创建摘要工作表"""
    summary_sheet = workbook.create_sheet("Summary", 0)
    
    # 统计信息
    total_matches = len(comparison_results)
    matches_with_candidates = sum(1 for r in comparison_results if r["ece_candidates"])
    
    # 添加摘要数据
    summary_data = [
        ["统计项", "数值"],
        ["总匹配项数", total_matches],
        ["有候选项的匹配数", matches_with_candidates],
        ["", ""],
        ["说明", ""],
        ["蓝色背景", "GSO章节信息"],
        ["橙色背景", "Ground Truth列（用于填写正确匹配编号）"],
        ["绿色背景", "ECE候选项信息"],
        ["Parameters", "保留完整的JSON结构化信息"]
    ]
    
    for row in summary_data:
        summary_sheet.append(row)
    
    # 设置列宽
    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 15
    
    return summary_sheet

def main():
    """主函数"""
    print("开始创建Excel报告...")
    
    try:
        # 文件路径
        input_file = r"D:/汽车检测知识图谱Agent/汽车检测知识图谱Agent/示例文件/temp/CELEX_42006X1227(06)_EN_TXT/comparison_results.json"
        output_file = r"D:/汽车检测知识图谱Agent/汽车检测知识图谱Agent/示例文件/temp/CELEX_42006X1227(06)_EN_TXT/comparison.xlsx"
        
        # 检查输入文件是否存在
        if not os.path.exists(input_file):
            print(f"错误: 找不到输入文件 {input_file}")
            return
        
        # 加载数据
        comparison_results = load_comparison_results(input_file)
        print(f"加载了 {len(comparison_results)} 个对比结果")
        
        # 创建Excel报告
        create_excel_report(comparison_results, output_file)
        
        # 创建带摘要的增强版本
        workbook = Workbook()
        
        # 删除默认工作表
        workbook.remove(workbook.active)
        
        # 创建摘要工作表
        create_summary_sheet(workbook, comparison_results)
        
        # 准备数据并创建主工作表
        excel_data = prepare_excel_data(comparison_results)
        df = pd.DataFrame(excel_data)
        
        main_sheet = workbook.create_sheet("Detailed Comparison")
        for r in dataframe_to_rows(df, index=False, header=True):
            main_sheet.append(r)
        
        # 设置样式
        style_excel_workbook(workbook, main_sheet, len(excel_data))
        merge_gso_cells(main_sheet, excel_data)
        main_sheet.freeze_panes = "A2"
        
        # 保存增强版本
        enhanced_output = output_file.replace('.xlsx', '_Enhanced.xlsx')
        workbook.save(enhanced_output)
        
        print(f"增强版Excel报告已保存到: {enhanced_output}")
        print("\n报告包含以下功能:")
        print("- 摘要工作表：统计信息和说明")
        print("- 详细对比工作表：完整的匹配对比数据")
        print("- Ground Truth列：用于填写正确匹配的候选项编号（橙色背景）")
        print("- 颜色编码：蓝色=GSO信息，橙色=Ground Truth，绿色=ECE候选项")
        print("- Parameters保留完整JSON结构")
        print("- 单元格合并：GSO章节信息自动合并")
        print("- 冻结窗格：便于浏览大量数据")
        
    except Exception as e:
        print(f"创建Excel报告时出现错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
